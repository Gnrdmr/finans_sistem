from datetime import date
from decimal import Decimal
import openpyxl
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib import messages
from django.http import HttpResponse
from dateutil.relativedelta import relativedelta
from datetime import timedelta
from collections import defaultdict
from django.core.paginator import Paginator
from .forms import TransactionForm, BudgetLimitForm, RecurringTransactionForm
from .utils import convert_to_try, get_exchange_rates 
from .forms import ExpenseGroupForm, SharedExpenseForm, TransactionTemplateForm
from .models import ExpenseGroup
from .models import Category
from tracker.models import Transaction, BudgetLimit, RecurringTransaction
from .models import Transaction, TransactionTemplate, BudgetLimit, RecurringTransaction, ExpenseGroup, SharedExpense
from .forms import SubscriptionForm
from django.utils import timezone
from .models import SavingsProfile
from .forms import SavingsProfileForm
from decimal import Decimal
from .models import BudgetLimit
from django.db.models import Sum
from datetime import datetime
from .models import CreditCard
from .forms import CreditCardForm
from .models import SavingsGoal
from .forms import SavingsGoalForm



# 1. Kayıt Olma Görünümü
def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'tracker/register.html', {'form': form})


# 2. Giriş Yapma Görünümü
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'tracker/login.html', {'form': form})


# 3. Çıkış Yapma Görünümü
def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')
    return render(request, 'tracker/logout.html')


def delete_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    transaction.delete()
    return redirect('home')


# 4. Ana Sayfa, Tekrarlayan İşlem Tetikleyicisi, Akıllı Bütçe Uyarıları, TRY Özet Hesabı, Canlı Kurlar, Finansal Sağlık ve Grafik Verileri (Gün 6-13)
# 4. Ana Sayfa, Tekrarlayan İşlem Tetikleyicisi, Akıllı Bütçe Uyarıları, TRY Özet Hesabı, Canlı Kurlar, Finansal Sağlık ve Grafik Verileri (Gün 6-13)
@login_required(login_url='login')
def home(request):
    user_transactions = Transaction.objects.filter(user=request.user)
    budgets = BudgetLimit.objects.filter(user=request.user)
    user_limits = BudgetLimit.objects.filter(user=request.user)
    


    search_query = request.GET.get('q', '')
    category_filter = request.GET.get('category', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')

    if search_query:
        user_transactions = user_transactions.filter(title__icontains=search_query)
    if category_filter:
        user_transactions = user_transactions.filter(category_id=category_filter)
    if start_date:
        user_transactions = user_transactions.filter(date__gte=start_date)
    if end_date:
        user_transactions = user_transactions.filter(date__lte=end_date)
    if min_amount:
        user_transactions = user_transactions.filter(amount__gte=min_amount)
    if max_amount:
        user_transactions = user_transactions.filter(amount__lte=max_amount)

        
    # --- GÜN 15: Sayfalama (Pagination) Entegrasyonu (Sayfa başına 10 kayıt) ---
    paginator = Paginator(user_transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # --- B. Tekrarlayan İşlemler Otomatik Kontrolü (Gün 7) ---
    today = date.today()
    recurring_items = RecurringTransaction.objects.filter(user=request.user, is_active=True, next_date__lte=today)
    
    for item in recurring_items:
        Transaction.objects.create(
            user=item.user,
            title=f"{item.title} (Tekrarlayan)",
            amount=item.amount,
            transaction_type=item.transaction_type,
            date=item.next_date,
            currency=item.currency,
            category=item.category,
            description=f"Otomatik oluşturulan tekrarlayan işlem ({item.get_interval_display()})"
        )
        
        if item.interval == 'DAILY':
            item.next_date += timedelta(days=1)
        elif item.interval == 'WEEKLY':
            item.next_date += timedelta(weeks=1)
        elif item.interval == 'MONTHLY':
            item.next_date += relativedelta(months=1)
        elif item.interval == 'YEARLY':
            item.next_date += relativedelta(years=1)
            
        item.save()
        messages.info(request, f"🔄 Tekrarlayan işlem otomatik eklendi: {item.title} ({item.amount} {item.currency})")

    # --- C. AKILLI BÜTÇE UYARI ALGORİTMASI (%80 - %100) (Gün 6) ---
    current_month = date.today().month
    current_year = date.today().year
    
    for limit_obj in user_limits:
        category_expenses = user_transactions.filter(
            category=limit_obj.category,
            transaction_type='EXPENSE',
            date__year=current_year,
            date__month=current_month
        )
        
        total_spent = sum([t.try_amount for t in category_expenses], Decimal('0.00'))
        limit_val = limit_obj.monthly_limit
        
        if limit_val > 0:
            percentage = (total_spent / limit_val) * Decimal('100')
            
            if percentage >= 100:
                messages.error(request, f"🚨 DİKKAT! '{limit_obj.category.name}' kategorisindeki aylık bütçe limitinizi aştınız! Harcama: {total_spent} TRY / Limit: {limit_val} TRY")
            elif percentage >= 80:
                messages.warning(request, f"⚠️ UYARI: '{limit_obj.category.name}' kategorisindeki bütçenizin %80'ine ulaştınız. Harcama: {total_spent} TRY")

    # --- D. TOPLAM GELİR VE GİDERLERİN TRY KARŞILIĞI HESABI (Gün 8) ---
    total_income_try = Decimal('0.00')
    total_expense_try = Decimal('0.00')
    
    for t in user_transactions:

        t.try_amount = convert_to_try(t.amount, t.currency)

        if t.transaction_type == 'Gelir' or t.transaction_type == 'INCOME':
            total_income_try += t.try_amount
        else:
            total_expense_try += t.try_amount

    # --- D-2. GÜN 12: NET NAKİT AKIŞI VE TASARRUF ORANI HESABI ---
    net_cash_flow = total_income_try - total_expense_try
    
    if total_income_try > 0:
        savings_rate = (net_cash_flow / total_income_try) * 100
    else:
        savings_rate = Decimal('0.00')

    # --- E. CANLI KURLARI ÇEKME VE CONTEXT'E EKLEME (Gün 8) ---
    rates = get_exchange_rates()
    usd_rate = rates.get('USD', 0)
    eur_rate = rates.get('EUR', 0)

    
    category_chart_data = {}
    expense_transactions = user_transactions.filter(transaction_type='EXPENSE')
    
    for t in expense_transactions:
        cat_name = t.category.name if t.category else "Diğer"
        
        amt = float(convert_to_try(t.amount, t.currency))
        category_chart_data[cat_name] = category_chart_data.get(cat_name, 0.0) + amt

    category_labels = list(category_chart_data.keys())
    category_values = list(category_chart_data.values())


    monthly_income = defaultdict(float)
    monthly_expense = defaultdict(float)
    
    for t in user_transactions:
        
        month_key = t.date.strftime('%Y-%m')
        amt = float(t.try_amount)
        
        if t.transaction_type in ['Gelir', 'INCOME']:
            monthly_income[month_key] += amt
        else:
            monthly_expense[month_key] += amt
            
    
    sorted_months = sorted(list(set(list(monthly_income.keys()) + list(monthly_expense.keys()))))
    
    trend_labels = sorted_months
    trend_income_values = [monthly_income[m] for m in sorted_months]
    trend_expense_values = [monthly_expense[m] for m in sorted_months]

    context = {
        'transactions': user_transactions,  # Normal user_transactions yerine sayfalama objesini gönderiyoruz
        'page_obj': page_obj,
        'user_limits': user_limits,
        'categories': Category.objects.all(), # Filtre formu için kategoriler
        'total_income_try': total_income_try,
        'total_expense_try': total_expense_try,
        'net_cash_flow': net_cash_flow,
        'savings_rate': savings_rate,
        'usd_rate': usd_rate,
        'eur_rate': eur_rate,
        'category_labels_json': json.dumps(category_labels),
        'category_values_json': json.dumps(category_values),
        'trend_labels_json': json.dumps(trend_labels),
        'trend_income_json': json.dumps(trend_income_values),
        'trend_expense_json': json.dumps(trend_expense_values),
        'budgets': budgets,
        'user_limits': user_limits,
        

  } 
    return render(request, 'tracker/home.html', context)


@login_required(login_url='login')
def transaction_add(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
        
            
            return redirect('home') 
    else:
        form = TransactionForm()
        
    return render(request, 'tracker/transaction_form.html', {'form': form, 'action': 'Ekle'})

# 6. İşlem Düzenleme (Update)
@login_required(login_url='login')
def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            messages.success(request, "İşlem başarıyla güncellendi!")
            return redirect('home')
    else:
        form = TransactionForm(instance=transaction)
        
    return render(request, 'tracker/transaction_form.html', {'form': form, 'action': 'Düzenle'})


# 7. İşlem Silme (Delete)
@login_required(login_url='login')
def transaction_delete(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    
    if request.method == 'POST':
        transaction.delete()
        messages.error(request, "İşlem sistemden silindi.")
        return redirect('home')
        
    return render(request, 'tracker/transaction_confirm_delete.html', {'transaction': transaction})


# 8. Excel Dışa Aktarma (Export)
@login_required(login_url='login')
def export_transactions_excel(request):
    transactions = Transaction.objects.filter(user=request.user)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finansal Islemler"
    
    ws.append(["Baslik", "Miktar", "Tur", "Tarih", "Doviz", "Aciklama"])
    
    for t in transactions:
        ws.append([
            t.title,
            float(t.amount),
            t.get_transaction_type_display(),
            str(t.date),
            t.currency,
            t.description or ""
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=finans_raporu.xlsx'
    wb.save(response)
    
    return response


# 9. Excel İçe Aktarma (Import)
@login_required(login_url='login')
def import_transactions_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                title, amount, trans_type_display, date_val, currency, description = row[:6]
                
                if not title or not amount:
                    continue
                
                trans_type = 'INCOME' if trans_type_display == 'Gelir' else 'EXPENSE'
                
                Transaction.objects.create(
                    user=request.user,
                    title=title,
                    amount=amount,
                    transaction_type=trans_type,
                    date=date_val if date_val else date.today(),
                    currency=currency if currency else 'TRY',
                    description=description
                )
                
            messages.success(request, "Excel dosyasındaki veriler başarıyla içe aktarıldı!")
        except Exception as e:
            messages.error(request, f"Dosya okunurken bir hata oluştu: {e}")
            
        return redirect('home')
        
    return render(request, 'tracker/import_excel.html')


# 10. Kategori Bütçe Limiti Belirleme Görünümü (Gün 6)
@login_required(login_url='login')
def set_budget_limit(request):
    if request.method == 'POST':
        form = BudgetLimitForm(request.POST)
        if form.is_valid():
            limit_obj = form.save(commit=False)
            limit_obj.user = request.user
            
            existing = BudgetLimit.objects.filter(user=request.user, category=limit_obj.category).first()
            if existing:
                existing.monthly_limit = limit_obj.monthly_limit
                existing.save()
            else:
                limit_obj.save()
                
            messages.success(request, "Bütçe limiti başarıyla kaydedildi!")
            return redirect('home')
    else:
        form = BudgetLimitForm()
        
    return render(request, 'tracker/set_limit.html', {'form': form})


# 11. Tekrarlayan İşlem Tanımlama Görünümü (Gün 7)
@login_required(login_url='login')
def recurring_add(request):
    if request.method == 'POST':
        form = RecurringTransactionForm(request.POST)
        if form.is_valid():
            rec_item = form.save(commit=False)
            rec_item.user = request.user
            rec_item.save()
            messages.success(request, "Tekrarlayan işlem başarıyla tanımlandı!")
            return redirect('home')
    else:
        form = RecurringTransactionForm()
        
    return render(request, 'tracker/recurring_form.html', {'form': form})


@login_required(login_url='login')
def create_expense_group(request):
    if request.method == 'POST':
        form = ExpenseGroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)
            group.created_by = request.user
            group.save()
            form.save_m2m() 
            messages.success(request, "Ortak harcama grubu başarıyla oluşturuldu!")
            return redirect('home')
    else:
        form = ExpenseGroupForm()
    return render(request, 'tracker/group_form.html', {'form': form})



@login_required(login_url='login')
def add_shared_expense(request):
    if request.method == 'POST':
        form = SharedExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ortak harcama sisteme başarıyla eklendi!")
            return redirect('home')
    else:
        form = SharedExpenseForm()
    return render(request, 'tracker/shared_expense_form.html', {'form': form})


@login_required(login_url='login')
def group_list(request):
    user_groups = ExpenseGroup.objects.filter(members=request.user).distinct()
    return render(request, 'tracker/group_list.html', {'user_groups': user_groups})



@login_required(login_url='login')
def group_detail(request, group_id):
    group = get_object_or_404(ExpenseGroup, pk=group_id, members=request.user)
    expenses = group.expenses.all()
    members = group.members.all()
    member_count = members.count()
    
    
    total_spent = sum([convert_to_try(exp.amount, exp.currency) for exp in expenses])
    
    
    per_person_share = total_spent / member_count if member_count > 0 else Decimal('0.00')
    
    
    member_balances = {}
    for member in members:
        user_paid = sum([
            convert_to_try(exp.amount, exp.currency) 
            for exp in expenses if exp.paid_by == member
        ])
        
        net_balance = user_paid - per_person_share
        member_balances[member] = {
            'paid': user_paid,
            'balance': net_balance
        }

    context = {
        'group': group,
        'expenses': expenses,
        'total_spent': total_spent,
        'per_person_share': per_person_share,
        'member_balances': member_balances,
    }
    return render(request, 'tracker/group_detail.html', context)



# 16. Grup Borçlarını Kapatma / Sıfırlama Görünümü (Gün 10)
@login_required(login_url='login')
def settle_debts(request, group_id):
    group = get_object_or_404(ExpenseGroup, pk=group_id, members=request.user)
    if request.method == 'POST':
        # Gruptaki tüm harcamaları silerek borçları sıfırlıyoruz (netleştirme)
        group.expenses.all().delete()
        messages.success(request, f"'{group.name}' grubundaki tüm borçlar başarıyla kapatıldı ve hesaplar netleştirildi!")
    return redirect('group_detail', group_id=group.id)





@login_required(login_url='login')
def monthly_analytics(request):
    today = date.today()
    current_year_month = (today.year, today.month)
    

    current_month_start = date(today.year, today.month, 1)
    if today.month == 1:
        last_month_start = date(today.year - 1, 12, 1)
        last_month_end = date(today.year - 1, 12, 31)
    else:
        last_month_start = date(today.year, today.month - 1, 1)
        last_month_end = current_month_start - relativedelta(days=1)

    
    user_transactions = Transaction.objects.filter(user=request.user, transaction_type='Gider')

    
    current_expenses = user_transactions.filter(date__gte=current_month_start, date__lte=today)
    current_categories = {}
    for t in current_expenses:
        cat_name = t.category.name if t.category else "Diğer"
        amt = convert_to_try(t.amount, t.currency)
        current_categories[cat_name] = current_categories.get(cat_name, 0) + amt


    last_expenses = user_transactions.filter(date__gte=last_month_start, date__lte=last_month_end)
    last_categories = {}
    for t in last_expenses:
        cat_name = t.category.name if t.category else "Diğer"
        amt = convert_to_try(t.amount, t.currency)
        last_categories[cat_name] = last_categories.get(cat_name, 0) + amt

    
    insights = []
    all_categories = set(list(current_categories.keys()) + list(last_categories.keys()))

    for cat in all_categories:
        curr_val = current_categories.get(cat, 0)
        last_val = last_categories.get(cat, 0)

        if last_val > 0:
            change_percent = ((curr_val - last_val) / last_val) * 100
            if change_percent > 0:
                insights.append({
                    'category': cat,
                    'text': f"Bu ay {cat} harcamaların geçen aya göre %{change_percent:.1f} arttı.",
                    'type': 'danger' # Artış gider için genelde olmuyor (kırmızı)
                })
            elif change_percent < 0:
                saving_percent = abs(change_percent)
                insights.append({
                    'category': cat,
                    'text': f"{cat} harcamalarında geçen aya göre %{saving_percent:.1f} tasarruf ettin!",
                    'type': 'success' # Tasarruf yeşil
                })
        elif curr_val > 0:
            insights.append({
                'category': cat,
                'text': f"Geçen ay hiç yapılmayan {cat} kategorisinde bu ay {curr_val:.2f} TRY harcama yapıldı.",
                'type': 'info'
            })

    context = {
        'insights': insights,
        'current_categories': current_categories,
        'last_categories': last_categories,
    }
    return render(request, 'tracker/monthly_analytics.html', context)



def subscription_calendar_view(request):
    today = date.today()
    
    
    upcoming_bills = Transaction.objects.filter(
        user=request.user,
        is_subscription=True,
        is_paid=False,
        due_date__gte=today 

    ).order_by('due_date')

    context = {
        'upcoming_bills': upcoming_bills,
    }
    return render(request, 'tracker/subscription_calendar.html', context)





def subscription_calendar_view(request):
    today = date.today()
    
    if request.method == 'POST':
        form = SubscriptionForm(request.POST)
        if form.is_valid():
            subscription = form.save(commit=False)
            subscription.user = request.user
            subscription.is_subscription = True  
            subscription.is_paid = False
            subscription.save()
            return redirect('subscription_calendar')
    else:
        form = SubscriptionForm()

    upcoming_bills = Transaction.objects.filter(
        user=request.user,
        is_subscription=True,
        is_paid=False,
        due_date__gte=today
    ).order_by('due_date')

    context = {
        'upcoming_bills': upcoming_bills,
        'form': form,
    }
    return render(request, 'tracker/subscription_calendar.html', context)


def templates_management_view(request):
    """Kullanıcının şablonlarını gördüğü ve yeni şablon eklediği sayfa"""
    if request.method == 'POST':
        form = TransactionTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.user = request.user
            template.save()
            return redirect('manage_templates')
    else:
        form = TransactionTemplateForm()

    user_templates = TransactionTemplate.objects.filter(user=request.user)

    context = {
        'form': form,
        'user_templates': user_templates,
    }
    return render(request, 'tracker/manage_templates.html', context)

def quick_add_transaction_view(request, template_id):
    """Şablon butonuna tıklandığında anında işlem (harcama) oluşturan fonksiyon"""
    template = get_object_or_404(TransactionTemplate, id=template_id, user=request.user)
    
    
    Transaction.objects.create(
        user=request.user,
        title=template.title,
        amount=template.amount,
        date=timezone.now().date(),
        
    )
    
    return redirect('home') 

def savings_simulation_view(request):
    profile, created = SavingsProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = SavingsProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('savings_simulation')
    else:
        form = SavingsProfileForm(instance=profile)

    
    from decimal import Decimal



    total_income = sum(t.amount for t in Transaction.objects.filter(user=request.user, amount__gt=0))

    
    inv_amount = (total_income * Decimal(str(profile.investment_rate))) / Decimal('100')
    em_amount  = (total_income * Decimal(str(profile.emergency_rate))) / Decimal('100')
    oth_amount = (total_income * Decimal(str(profile.other_rate))) / Decimal('100')

    context = {
        'form': form,
        'profile': profile,
        'total_income': total_income,
        'inv_amount': inv_amount,
        'em_amount': em_amount,
        'oth_amount': oth_amount,
    }
    return render(request, 'tracker/savings_simulation.html', context)



@login_required(login_url='login')
def credit_cards_view(request):
    cards = CreditCard.objects.filter(user=request.user)
    form = CreditCardForm()
    
    if request.method == 'POST':
        form = CreditCardForm(request.POST)
        if form.is_valid():
            card = form.save(commit=False)
            card.user = request.user
            card.save()
            messages.success(request, "Kredi kartı başarıyla eklendi!")
            return redirect('credit_cards')
            
    context = {
        'cards': cards,
        'form': form,
    }
    return render(request, 'tracker/credit_cards.html', context)   




@login_required(login_url='login')
def savings_list(request):
    goals = SavingsGoal.objects.filter(user=request.user)
    form = SavingsGoalForm()
    
    if request.method == 'POST':
        form = SavingsGoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user
            goal.save()
            messages.success(request, "Yeni tasarruf hedefi başarıyla eklendi!")
            return redirect('savings_list')
            
    context = {
        'goals': goals,
        'form': form,
    }
    return render(request, 'tracker/savings_list.html', context)







